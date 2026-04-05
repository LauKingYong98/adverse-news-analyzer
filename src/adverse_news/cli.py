import argparse
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook

from adverse_news.analyzer import analyze_company
from adverse_news.models import CompanyInput
from adverse_news.report.excel_writer import write_report


def _read_companies_from_excel(path: str) -> list[CompanyInput]:
    """Read company names from column A, aliases from column B."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    companies = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        name = row[0]
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        aliases = []
        if len(row) > 1 and row[1]:
            aliases = [a.strip() for a in str(row[1]).split(",") if a.strip()]
        companies.append(CompanyInput(name=name, aliases=aliases))
    wb.close()
    return companies


def _run_single(company: CompanyInput, output_path: str, api_key: str | None, model: str | None = None):
    print(f"\n  Analyzing: {company.name}...")
    report = analyze_company(company, api_key=api_key, model=model)
    write_report(report, output_path)

    print(f"  Total articles:  {report.total_articles_found}")
    print(f"  Positive: {len(report.positive_results)} | "
          f"Negative: {len(report.negative_results)} | "
          f"Neutral: {len(report.neutral_results)}")
    if report.top_risk_factors:
        print(f"  Top risks: {', '.join(report.top_risk_factors[:5])}")
    print(f"  Report: {output_path}")
    return report


def main():
    parser = argparse.ArgumentParser(
        description="Adverse News Analyzer — screen companies for negative news impacting fair valuation"
    )
    parser.add_argument("--company", "-c", default=None, help="Company name to analyze")
    parser.add_argument("--aliases", "-a", default="", help="Comma-separated aliases/short forms")
    parser.add_argument(
        "--input-excel", "-i", default=None,
        help="Excel file with company list (col A: name, col B: aliases)"
    )
    parser.add_argument("--months", "-m", type=int, default=12, help="Search period in months (default: 12)")
    parser.add_argument(
        "--output", "-o", default=None,
        help="Output path. Single: file path. Batch: directory (default: ./reports/)"
    )
    parser.add_argument("--api-key", default=None, help="Anthropic API key (or set ANTHROPIC_API_KEY env var)")
    parser.add_argument(
        "--model", default=None,
        help="LLM model. claude-* uses API; others use Ollama (default: claude-sonnet-4-6)"
    )
    parser.add_argument("--verbose", "-v", action="store_true", help="Enable verbose logging")

    args = parser.parse_args()

    if not args.company and not args.input_excel:
        parser.error("Either --company or --input-excel is required")

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )

    print(f"\n{'='*60}")
    print(f"  Adverse News Analyzer")
    print(f"{'='*60}")

    if args.input_excel:
        companies = _read_companies_from_excel(args.input_excel)
        if not companies:
            print("  No companies found in input file.")
            sys.exit(1)

        for c in companies:
            c.search_period_months = args.months

        output_dir = Path(args.output or "./reports")
        output_dir.mkdir(parents=True, exist_ok=True)

        print(f"  Mode:    Batch ({len(companies)} companies)")
        print(f"  Input:   {args.input_excel}")
        print(f"  Period:  {args.months} months")
        print(f"  Output:  {output_dir}/")
        print(f"{'='*60}")

        failed = []
        for idx, company in enumerate(companies, 1):
            print(f"\n  [{idx}/{len(companies)}] {company.name}")
            output_path = output_dir / f"{company.name.replace(' ', '_')}_report.xlsx"
            try:
                _run_single(company, str(output_path), args.api_key, args.model)
            except Exception as e:
                logging.error(f"Failed: {company.name}: {e}")
                failed.append(company.name)

        print(f"\n{'='*60}")
        print(f"  Done: {len(companies) - len(failed)}/{len(companies)} succeeded")
        if failed:
            print(f"  Failed: {', '.join(failed)}")
        print(f"{'='*60}\n")

    else:
        aliases = [a.strip() for a in args.aliases.split(",") if a.strip()] if args.aliases else []
        company = CompanyInput(name=args.company, aliases=aliases, search_period_months=args.months)
        output_path = args.output or f"{company.name.replace(' ', '_')}_report.xlsx"

        print(f"  Company: {company.name}")
        if aliases:
            print(f"  Aliases: {', '.join(aliases)}")
        print(f"  Period:  {args.months} months")
        print(f"  Output:  {output_path}")
        print(f"{'='*60}")

        try:
            _run_single(company, output_path, args.api_key, args.model)
            print(f"\n{'='*60}")
            print(f"  Done!")
            print(f"{'='*60}\n")
        except Exception as e:
            logging.error(f"Analysis failed: {e}", exc_info=True)
            sys.exit(1)


if __name__ == "__main__":
    main()
