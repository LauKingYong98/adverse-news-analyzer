import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from adverse_news.models import AnalysisReport, Sentiment

logger = logging.getLogger(__name__)

# Style constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11)
LINK_FONT = Font(color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SENTIMENT_FILLS = {
    Sentiment.POSITIVE: POSITIVE_FILL,
    Sentiment.NEGATIVE: NEGATIVE_FILL,
    Sentiment.NEUTRAL: NEUTRAL_FILL,
}


def _style_header_row(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _auto_width(ws, min_width: int = 10, max_width: int = 50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def write_report(report: AnalysisReport, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    wb = Workbook()

    # === Sheet 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.cell(row=1, column=1, value="Adverse News Analysis Report").font = TITLE_FONT
    ws_summary.merge_cells("A1:D1")

    rows = [
        ("Company", report.company.name),
        ("Aliases", ", ".join(report.company.aliases) if report.company.aliases else "N/A"),
        ("Search Period", f"{report.company.search_period_months} months"),
        ("Report Date", report.run_timestamp.strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total Articles Analyzed", report.total_articles_found),
        ("Positive", len(report.positive_results)),
        ("Negative", len(report.negative_results)),
        ("Neutral", len(report.neutral_results)),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws_summary.cell(row=i, column=1, value=label).font = SUBTITLE_FONT
        ws_summary.cell(row=i, column=2, value=str(value))

    # Top risk factors
    risk_row = len(rows) + 5
    ws_summary.cell(row=risk_row, column=1, value="Top Risk Factors").font = SUBTITLE_FONT
    if report.top_risk_factors:
        for j, factor in enumerate(report.top_risk_factors):
            ws_summary.cell(row=risk_row + 1 + j, column=1, value=f"  {j + 1}. {factor}")
    else:
        ws_summary.cell(row=risk_row + 1, column=1, value="  No significant risk factors identified")

    _auto_width(ws_summary)

    # === Sheet 2: All Articles ===
    ws_all = wb.create_sheet("All Articles")
    headers = ["Date", "Source", "Title", "URL", "Sentiment", "Confidence", "Summary", "Risk Factors"]
    for col, h in enumerate(headers, 1):
        ws_all.cell(row=1, column=col, value=h)
    _style_header_row(ws_all, len(headers))

    sorted_results = sorted(
        report.all_results,
        key=lambda r: r.article.published_date or datetime.min,
        reverse=True,
    )

    for row_idx, r in enumerate(sorted_results, start=2):
        date_str = r.article.published_date.strftime("%Y-%m-%d") if r.article.published_date else ""
        ws_all.cell(row=row_idx, column=1, value=date_str)
        ws_all.cell(row=row_idx, column=2, value=r.article.source)
        ws_all.cell(row=row_idx, column=3, value=r.article.title)

        url_cell = ws_all.cell(row=row_idx, column=4, value=r.article.url)
        url_cell.hyperlink = r.article.url
        url_cell.font = LINK_FONT

        sentiment_cell = ws_all.cell(row=row_idx, column=5, value=r.sentiment.value)
        sentiment_cell.fill = SENTIMENT_FILLS.get(r.sentiment, NEUTRAL_FILL)
        sentiment_cell.alignment = Alignment(horizontal="center")

        ws_all.cell(row=row_idx, column=6, value=round(r.confidence, 2))
        ws_all.cell(row=row_idx, column=7, value=r.summary)
        ws_all.cell(row=row_idx, column=8, value="; ".join(r.risk_factors) if r.risk_factors else "")

        for col in range(1, len(headers) + 1):
            ws_all.cell(row=row_idx, column=col).border = THIN_BORDER
            ws_all.cell(row=row_idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    _auto_width(ws_all)

    # === Sheet 3: Negative Events ===
    ws_neg = wb.create_sheet("Negative Events")
    neg_headers = ["Date", "Title", "Source", "Summary", "Risk Factors", "URL"]
    for col, h in enumerate(neg_headers, 1):
        ws_neg.cell(row=1, column=col, value=h)
    _style_header_row(ws_neg, len(neg_headers))

    neg_sorted = sorted(
        report.negative_results,
        key=lambda r: r.article.published_date or datetime.min,
        reverse=True,
    )

    for row_idx, r in enumerate(neg_sorted, start=2):
        date_str = r.article.published_date.strftime("%Y-%m-%d") if r.article.published_date else ""
        ws_neg.cell(row=row_idx, column=1, value=date_str)
        ws_neg.cell(row=row_idx, column=2, value=r.article.title)
        ws_neg.cell(row=row_idx, column=3, value=r.article.source)
        ws_neg.cell(row=row_idx, column=4, value=r.summary)
        ws_neg.cell(row=row_idx, column=5, value="; ".join(r.risk_factors) if r.risk_factors else "")

        url_cell = ws_neg.cell(row=row_idx, column=6, value=r.article.url)
        url_cell.hyperlink = r.article.url
        url_cell.font = LINK_FONT

        for col in range(1, len(neg_headers) + 1):
            ws_neg.cell(row=row_idx, column=col).border = THIN_BORDER
            ws_neg.cell(row=row_idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    _auto_width(ws_neg)

    # Footer
    footer_row = len(neg_sorted) + 3
    ws_neg.cell(
        row=footer_row,
        column=1,
        value=f"Generated: {report.run_timestamp.strftime('%Y-%m-%d %H:%M')} | "
        "AI-generated report — should be reviewed by a qualified auditor.",
    ).font = Font(italic=True, color="808080")

    wb.save(output_path)
    logger.info(f"Report saved to {output_path}")
    return output_path
