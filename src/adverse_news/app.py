import io
import logging
import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from adverse_news.analyzer import analyze_company
from adverse_news.models import AnalysisReport, CompanyInput
from adverse_news.report.excel_writer import write_report

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)

st.set_page_config(page_title="Adverse News Analyzer", page_icon="📊", layout="wide")

CLAUDE_MODELS = [
    "claude-sonnet-4-6",
    "claude-haiku-4-5-20251001",
    "claude-opus-4-6",
]

DEFAULT_OLLAMA_MODELS = [
    "llama3.1",
    "llama3.1:70b",
    "gemma3",
    "gemma4:12b",
    "gemma4:27b",
    "qwen2.5",
    "qwen2.5:32b",
    "qwen3",
    "qwen3:32b",
    "mistral",
    "deepseek-r1",
]


def _get_ollama_models() -> list[str]:
    try:
        import ollama
        models = ollama.list()
        names = [m.model for m in models.models] if models.models else []
        return names if names else DEFAULT_OLLAMA_MODELS
    except Exception:
        return DEFAULT_OLLAMA_MODELS


def _report_to_dataframe(report: AnalysisReport) -> pd.DataFrame:
    rows = []
    for r in report.all_results:
        rows.append({
            "Date": r.article.published_date.strftime("%Y-%m-%d") if r.article.published_date else "",
            "Source": r.article.source,
            "Title": r.article.title,
            "URL": r.article.url,
            "Sentiment": r.sentiment.value,
            "Confidence": round(r.confidence, 2),
            "Summary": r.summary,
            "Risk Factors": "; ".join(r.risk_factors) if r.risk_factors else "",
        })
    return pd.DataFrame(rows)


def _report_to_excel_bytes(report: AnalysisReport) -> bytes:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        write_report(report, f.name)
        return Path(f.name).read_bytes()


def _color_sentiment(val: str) -> str:
    colors = {
        "POSITIVE": "background-color: #C6EFCE; color: #006100",
        "NEGATIVE": "background-color: #FFC7CE; color: #9C0006",
        "NEUTRAL": "background-color: #F2F2F2; color: #333333",
    }
    return colors.get(val, "")


def _read_companies_from_upload(uploaded_file) -> list[CompanyInput]:
    wb = load_workbook(io.BytesIO(uploaded_file.getvalue()), read_only=True)
    ws = wb.active
    companies = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        name = row[0]
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        aliases = []
        if len(row) > 1 and row[1]:
            aliases = [a.strip() for a in str(row[1]).split(",") if a.strip()]
        companies.append(CompanyInput(name=name, aliases=aliases))
    wb.close()
    return companies


# --- Sidebar ---
with st.sidebar:
    st.header("Settings")

    provider = st.radio("AI Provider", ["Claude (API)", "Ollama (Local)"], index=0)

    if provider == "Claude (API)":
        model = st.selectbox("Model", CLAUDE_MODELS, index=0)
        api_key = st.text_input("Anthropic API Key", type="password", help="Or set ANTHROPIC_API_KEY in .env")
    else:
        ollama_models = _get_ollama_models()
        model = st.selectbox(
            "Ollama Model", ollama_models, index=0,
            help="Run `ollama pull <model>` to download new ones.",
        )
        custom_model = st.text_input("Or enter custom model name", placeholder="e.g. phi3:medium")
        if custom_model.strip():
            model = custom_model.strip()
        api_key = None

    st.divider()
    months = st.slider("Search Period (months)", 3, 18, 12)
    languages = st.multiselect("Languages", ["en", "zh"], default=["en", "zh"])

    st.divider()
    st.caption("Adverse News Analyzer v0.2")
    st.caption("AI-generated results — verify with professional judgment.")

# --- Main ---
st.title("Adverse News Analyzer")
st.markdown("Screen companies for adverse news impacting fair valuation.")

tab_single, tab_batch = st.tabs(["Single Company", "Batch (Excel Upload)"])

# === Single Company ===
with tab_single:
    col1, col2 = st.columns([2, 1])
    with col1:
        company_name = st.text_input("Company Name", placeholder="e.g. ByteDance")
    with col2:
        aliases_str = st.text_input("Aliases (comma-separated)", placeholder="e.g. TikTok, Douyin")

    analyze_btn = st.button("Analyze", key="single_analyze", type="primary", use_container_width=True)

    if analyze_btn and company_name:
        aliases = [a.strip() for a in aliases_str.split(",") if a.strip()] if aliases_str else []
        company = CompanyInput(name=company_name, aliases=aliases, search_period_months=months)

        from adverse_news.config import settings
        settings.languages = languages

        progress_bar = st.progress(0, text="Starting...")

        def update_progress(step: str, detail: str):
            steps = {"Searching": 0.10, "Parsing": 0.30, "Classifying": 0.60, "Complete": 1.0}
            for prefix, pct in steps.items():
                if step.startswith(prefix):
                    progress_bar.progress(pct, text=f"{step}: {detail}")
                    break

        with st.spinner("Running analysis..."):
            try:
                report = analyze_company(
                    company,
                    api_key=api_key if provider == "Claude (API)" else None,
                    model=model,
                    progress_callback=update_progress,
                )
                progress_bar.progress(1.0, text="Complete!")
            except Exception as e:
                st.error(f"Analysis failed: {e}")
                st.stop()

        # Results
        st.subheader("Results")
        col_pos, col_neg, col_neu, col_total = st.columns(4)
        col_pos.metric("Positive", len(report.positive_results))
        col_neg.metric("Negative", len(report.negative_results))
        col_neu.metric("Neutral", len(report.neutral_results))
        col_total.metric("Total", report.total_articles_found)

        if report.top_risk_factors:
            st.subheader("Top Risk Factors")
            for i, factor in enumerate(report.top_risk_factors[:5], 1):
                st.markdown(f"**{i}.** {factor}")

        if report.all_results:
            st.subheader("Articles")
            df = _report_to_dataframe(report)
            styled = df.style.map(_color_sentiment, subset=["Sentiment"])
            st.dataframe(styled, hide_index=True, width=None)

            excel_bytes = _report_to_excel_bytes(report)
            st.download_button(
                label="Download Excel Report",
                data=excel_bytes,
                file_name=f"{company.name.replace(' ', '_')}_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        else:
            st.info("No articles found for this company.")

    elif analyze_btn:
        st.warning("Please enter a company name.")

# === Batch ===
with tab_batch:
    st.markdown("""
    Upload an Excel file with companies to analyze.

    **Format:** Column A = Company Name (A1 = header), Column B = Aliases (optional, comma-separated)
    """)

    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"], key="batch_upload")

    if uploaded_file:
        companies = _read_companies_from_upload(uploaded_file)
        st.success(f"Found {len(companies)} companies:")
        for c in companies:
            alias_str = f" ({', '.join(c.aliases)})" if c.aliases else ""
            st.markdown(f"- **{c.name}**{alias_str}")

        batch_btn = st.button("Analyze All", key="batch_analyze", type="primary", use_container_width=True)

        if batch_btn and companies:
            from adverse_news.config import settings
            settings.languages = languages

            all_reports: list[tuple[CompanyInput, AnalysisReport]] = []
            progress = st.progress(0, text="Starting batch...")

            for idx, company in enumerate(companies):
                company.search_period_months = months
                progress.progress(
                    idx / len(companies),
                    text=f"[{idx + 1}/{len(companies)}] {company.name}...",
                )
                try:
                    report = analyze_company(
                        company,
                        api_key=api_key if provider == "Claude (API)" else None,
                        model=model,
                    )
                    all_reports.append((company, report))
                except Exception as e:
                    st.warning(f"Failed: {company.name} — {e}")

            progress.progress(1.0, text="Batch complete!")

            # Summary
            st.subheader("Batch Summary")
            summary_rows = []
            for company, report in all_reports:
                summary_rows.append({
                    "Company": company.name,
                    "Total": report.total_articles_found,
                    "Positive": len(report.positive_results),
                    "Negative": len(report.negative_results),
                    "Neutral": len(report.neutral_results),
                    "Top Risk Factors": ", ".join(report.top_risk_factors[:3]) or "None",
                })
            st.dataframe(pd.DataFrame(summary_rows), hide_index=True, width=None)

            # Downloads
            st.subheader("Download Reports")
            for company, report in all_reports:
                if report.all_results:
                    excel_bytes = _report_to_excel_bytes(report)
                    st.download_button(
                        label=f"Download: {company.name}",
                        data=excel_bytes,
                        file_name=f"{company.name.replace(' ', '_')}_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_{company.name}",
                    )

            # Detail view
            st.subheader("Details")
            for company, report in all_reports:
                with st.expander(
                    f"{company.name} — {len(report.negative_results)} neg / "
                    f"{len(report.positive_results)} pos / {len(report.neutral_results)} neu"
                ):
                    if report.all_results:
                        df = _report_to_dataframe(report)
                        styled = df.style.map(_color_sentiment, subset=["Sentiment"])
                        st.dataframe(styled, hide_index=True, width=None)
                    else:
                        st.info("No articles found.")
